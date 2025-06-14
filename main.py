import openpyxl
import os
import requests

from docx import Document
from docxcompose.composer import Composer
from docxtpl import DocxTemplate
from dotenv import load_dotenv


def cleanup():
    for file in os.listdir("tmp"):
        if file:
            os.remove(os.path.join("tmp", file))

    os.remove("guest.xlsx")

def merge_documents(temp_path, output_path):
    master = None
    for index, file in enumerate(os.listdir(temp_path)):
        file = temp_path + "/" + file
        if index == 0:
            master = Document(file)
            continue
        composer = Composer(master)
        append = Document(file)
        composer.append(append)
        composer.save(output_path)


def generate_envelope(template_path, output_path, addresses):
    output_path = output_path + "/output_%d.docx"
    for index, address in enumerate(addresses):
        # Load the template
        doc = DocxTemplate(template_path)
        # Replace the placeholders with the provided address information
        doc.render(address, autoescape=True)
        # Save the modified document as a new file
        doc.save(output_path.replace("%d", str(index)))

def get_sheet():
    sheet_id = "1vyP000arby2FBn82CuRk9HX3Ai6Cy4O4ddV4ofa-Q0E"
    gid = "1553461875"
    url = "https://docs.google.com/spreadsheets/d/{}/export?format=xlsx&gid={}#gid={}".format(sheet_id, gid, gid)
    response = requests.get(url)
    with open("guest.xlsx", "wb") as f:
        f.write(response.content)
    return response.content


def get_int(num):
    try:
        return int(num)
    except Exception:
        return None


def iter_rows():
    wb = openpyxl.load_workbook("guest.xlsx")
    ws = wb.active

    data = []

    for i in range(5, ws.max_row):
        # Get all the cells in the row
        cells = [" " if cell.value is None else cell.value for cell in ws[i]]

        partner_ref_column = 14
        id_column = 15

        print_state = cells[0]
        prefix = cells[1]
        first_name = cells[2]
        last_name = cells[3]
        invited_by = cells[4]
        rehersal_dinner = cells[6]
        suffix = cells[8]
        address_one = cells[9]
        address_two = cells[10]
        city = cells[11]
        state = cells[12]
        zip_code = cells[13]
        partner_ref = get_int(cells[partner_ref_column])
        id = get_int(cells[15])
        completed = cells[16]
        ignore = cells[17]

        if print_state == "Print" and ignore == "F" and completed == "F":
            names = [str.join(" ", [prefix, first_name, last_name, suffix])]
            # If partner ref is found, members living together but unwed
            if partner_ref:
                for partner_row in ws.iter_rows(min_row=5):
                    if get_int(partner_row[id_column].value) == partner_ref:
                        names.append(str.join(" ", [partner_row[1].value, partner_row[2].value, partner_row[3].value]))
                        break

            data.append({
                "NAMES": names,
                "ADDRESS1": address_one,
                "ADDRESS2": address_two,
                "CITY": city,
                "STATE": state,
                "ZIP": zip_code
            })

    return data

dotenv_path = os.path.join(os.path.dirname(__file__), '.env')
load_dotenv(dotenv_path)
template_path = os.environ.get("TEMPLATE_PATH")
output_path = os.environ.get("OUTPUT_PATH")
temp_path = os.environ.get("TEMP_PATH")

try:
    cleanup()
    # Remove completed file is exist
    os.remove(os.path.normpath(output_path))
except Exception as e:
    print(e)
    pass


try:
    # Download Sheet
    get_sheet()
    # Parse Sheet for Addresses
    addresses = iter_rows()
    # Create a Word document with the address data based off the template
    generate_envelope(template_path, temp_path, addresses)
    # Merge all Word Documents into one
    merge_documents(temp_path, output_path)
    # Clean temp files
    cleanup()
except Exception as e:
    print(e.format_exc())
